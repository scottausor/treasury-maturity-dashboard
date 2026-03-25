#!/usr/bin/env python3
"""
Treasury MSPD Scraper
=====================
Reads the Monthly Statement of Public Debt (MSPD) Excel file published by the
US Treasury each month, extracts all outstanding marketable securities with
their maturity dates, aggregates month-by-month maturities, and produces:
  - A verification Excel workbook  (maturity_data_YYYYMM.xlsx)
  - A JSON data file               (maturity_data_latest.json)
  - A standalone HTML dashboard    (treasury_dashboard_YYYYMM.html)

Source file structure (sheet 'Marketable'):
  Col B (index 1)  — section markers, e.g. "Treasury Bills (Maturity Value):"
  Col H (index 7)  — maturity date
  Col P (index 15) — amount outstanding ($ millions)

Usage:
    python treasury_scraper.py                          # auto-detect latest month
    python treasury_scraper.py --year 2026 --month 2    # specific month
    python treasury_scraper.py --xls path/to/file.xls   # use local file
"""

import requests
import pandas as pd
import json
import logging
import argparse
import sys
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path

# ── Logging ───────────────────────────────────────────────────────────────────

LOG_FILE = Path(__file__).parent / "treasury_scraper.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────

BASE_URL = (
    "https://fiscaldata.treasury.gov/static-data/published-reports/"
    "mspd-entire/MonthlyStatementPublicDebt_Entire_{YYYYMM}.xls"
)

# Row ranges for each security type in the 'Marketable' sheet (0-indexed).
# These are stable across MSPD editions — the section headers in col B confirm
# them. If Treasury ever adds new securities the boundaries will shift; the
# validation step will catch unexpected totals.
SECTIONS = {
    "bill":  (6,   93),   # Treasury Bills
    "note":  (99,  419),  # Treasury Notes
    "bond":  (424, 723),  # Treasury Bonds
    "tips":  (729, 860),  # TIPS
    "frn":   (864, 887),  # Floating Rate Notes
}

# Expected section-total rows (col B label, col P total) used for validation
TOTAL_ROWS = {
    "bill":  93,
    "note":  419,
    "bond":  723,
    "tips":  860,
    "frn":   887,
}


# ── Download ──────────────────────────────────────────────────────────────────

def build_url(year: int, month: int) -> str:
    return BASE_URL.format(YYYYMM=f"{year}{month:02d}")


def get_candidate_months() -> list[tuple[int, int]]:
    """Try last 3 months — MSPD is published ~5th of the following month."""
    today = date.today()
    return [(( today - relativedelta(months=d)).year,
              (today - relativedelta(months=d)).month) for d in range(0, 4)]


def download_xls(url: str, save_path: Path) -> Path:
    log.info(f"Downloading: {url}")
    resp = requests.get(url, timeout=60, stream=True)
    resp.raise_for_status()
    save_path.parent.mkdir(parents=True, exist_ok=True)
    with open(save_path, "wb") as f:
        for chunk in resp.iter_content(65536):
            f.write(chunk)
    log.info(f"Saved {save_path.stat().st_size / 1_048_576:.1f} MB → {save_path}")
    return save_path


def find_or_download_xls(year: int = None, month: int = None) -> tuple[Path, str]:
    cache_dir = Path(__file__).parent / "xls_cache"
    cache_dir.mkdir(exist_ok=True)

    candidates = [(year, month)] if year and month else get_candidate_months()

    for y, m in candidates:
        label = f"{y}{m:02d}"
        cached = cache_dir / f"MSPD_{label}.xls"
        if cached.exists():
            log.info(f"Using cached file: {cached}")
            return cached, label
        url = build_url(y, m)
        try:
            r = requests.head(url, timeout=15)
            if r.status_code == 200:
                return download_xls(url, cached), label
            log.warning(f"HTTP {r.status_code} for {label}, trying earlier month…")
        except requests.RequestException as e:
            log.warning(f"Could not reach {url}: {e}")

    raise RuntimeError(
        "Could not find or download any MSPD file. "
        "Pass --xls with a local file path, or --year/--month to specify a period."
    )


# ── Extraction ────────────────────────────────────────────────────────────────

def extract_from_excel(xls_path: Path) -> pd.DataFrame:
    """
    Read the 'Marketable' sheet and extract one row per security.
    Returns a DataFrame with columns:
        type, maturity_date, amount_millions
    """
    log.info(f"Reading: {xls_path}")
    raw = pd.read_excel(xls_path, sheet_name="Marketable",
                        header=None, engine="xlrd")
    log.info(f"Sheet dimensions: {raw.shape[0]} rows × {raw.shape[1]} cols")

    records = []
    for sec_type, (start, end) in SECTIONS.items():
        section_count = 0
        for i in range(start, end):
            raw_date = raw.iat[i, 7]   # Col H — maturity date
            raw_amt  = raw.iat[i, 15]  # Col P — amount outstanding ($M)

            # Skip empty cells
            if pd.isna(raw_date) or pd.isna(raw_amt):
                continue
            # Skip non-date values like "Various" or dotted lines
            if not hasattr(raw_date, "year"):
                continue
            try:
                amt = float(raw_amt)
            except (ValueError, TypeError):
                continue
            if amt <= 0:
                continue

            records.append({
                "type":            sec_type,
                "maturity_date":   pd.Timestamp(raw_date),
                "amount_millions": amt,
            })
            section_count += 1

        log.info(f"  {sec_type.upper():5s}: {section_count} securities extracted")

    df = pd.DataFrame(records)
    log.info(f"Total extracted: {len(df)} securities, "
             f"${df['amount_millions'].sum() / 1e6:.2f}T outstanding")
    return df


# ── Validation ────────────────────────────────────────────────────────────────

def validate(df: pd.DataFrame, xls_path: Path) -> pd.DataFrame:
    """
    Cross-check extracted totals against the official subtotals in the sheet.
    Warns if any section is off by more than 0.01%.
    """
    raw = pd.read_excel(xls_path, sheet_name="Marketable",
                        header=None, engine="xlrd")

    log.info("Validating against sheet subtotals…")
    for sec_type, total_row in TOTAL_ROWS.items():
        official = raw.iat[total_row, 15]
        extracted = df[df["type"] == sec_type]["amount_millions"].sum()
        if pd.isna(official):
            continue
        try:
            official = float(official)
        except (ValueError, TypeError):
            continue
        diff_pct = abs(extracted - official) / official * 100
        status = "✓" if diff_pct < 0.01 else "⚠ MISMATCH"
        log.info(f"  {sec_type.upper():5s}: extracted ${extracted:,.0f}M "
                 f"vs official ${official:,.0f}M  ({diff_pct:.4f}%)  {status}")
        if diff_pct >= 0.01:
            log.warning(f"  {sec_type} total differs from sheet by {diff_pct:.4f}% "
                        f"— check SECTIONS row boundaries in scraper config.")
    return df


# ── Aggregation ───────────────────────────────────────────────────────────────

def aggregate_by_month(df: pd.DataFrame) -> dict:
    """Aggregate maturities by calendar month."""
    today = pd.Timestamp(date.today())
    future = df[df["maturity_date"] >= today].copy()
    future["ym"] = future["maturity_date"].dt.to_period("M")

    pivot = (
        future.groupby(["ym", "type"])["amount_millions"]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )
    pivot.columns.name = None
    pivot["total_millions"] = pivot.drop(columns="ym").sum(axis=1)
    pivot["ym"] = pivot["ym"].astype(str)
    pivot = pivot.sort_values("ym")

    monthly_records = pivot.to_dict(orient="records")

    # Next-12-month window for summary stats
    cutoff = (today + relativedelta(months=12)).strftime("%Y-%m")
    next12 = pivot[pivot["ym"] <= cutoff]

    total12 = next12["total_millions"].sum()
    peak_row = next12.loc[next12["total_millions"].idxmax()] if not next12.empty else {}

    summary = {
        "total_next_12m_billions":  round(float(total12) / 1000, 1),
        "peak_month":               str(peak_row.get("ym", "—")),
        "peak_month_billions":      round(float(peak_row.get("total_millions", 0)) / 1000, 1),
        "avg_monthly_billions":     round(float(next12["total_millions"].mean()) / 1000, 1)
                                    if not next12.empty else 0,
        "types_found":              sorted(df["type"].unique().tolist()),
        "extracted_at":             datetime.utcnow().isoformat() + "Z",
    }

    log.info(f"Next 12 months: ${summary['total_next_12m_billions']:.1f}B total")
    log.info(f"Peak month: {summary['peak_month']} (${summary['peak_month_billions']:.0f}B)")

    return {
        "monthly":   monthly_records,
        "summary":   summary,
        "raw_count": len(df),
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(raw_df: pd.DataFrame, agg: dict, output_path: Path) -> Path:
    """
    Two-sheet verification workbook:
      Sheet 1 — Raw Securities  : every security parsed from the MSPD
      Sheet 2 — Monthly Summary : aggregated maturities by month + type
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(name="Arial", bold=True, color="E6EDF3", size=10)
    BLUE_FONT   = Font(name="Arial", size=10, color="0000FF")
    BLACK_FONT  = Font(name="Arial", size=10, color="000000")
    ALT_FILL    = PatternFill("solid", fgColor="F9FAFB")
    TYPE_FILLS  = {
        "bill": "FFF3E0", "note": "E3F2FD", "bond": "E8F5E9",
        "tips": "FFEBEE", "frn":  "F3E5F5",
    }

    wb = Workbook()

    # ── Sheet 1: Raw Securities ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw Securities"

    headers   = ["Type", "Maturity Date", "Amount ($M)", "Amount ($B)"]
    col_widths = [12, 16, 16, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    df_sorted = raw_df.sort_values(["maturity_date", "type"]).reset_index(drop=True)
    for ri, row in enumerate(df_sorted.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor=TYPE_FILLS.get(row.type, "FFFFFF"))
        ws1.cell(ri, 1, row.type.upper()).font = BLUE_FONT
        ws1.cell(ri, 1).fill = fill
        ws1.cell(ri, 1).alignment = Alignment(horizontal="center")

        dc = ws1.cell(ri, 2, str(row.maturity_date)[:10])
        dc.font = BLUE_FONT; dc.fill = fill
        dc.alignment = Alignment(horizontal="center")

        ac = ws1.cell(ri, 3, row.amount_millions)
        ac.font = BLUE_FONT; ac.fill = fill
        ac.number_format = '"$"#,##0.0'; ac.alignment = Alignment(horizontal="right")

        bc = ws1.cell(ri, 4, f"=C{ri}/1000")
        bc.font = BLACK_FONT; bc.fill = fill
        bc.number_format = '"$"#,##0.00'; bc.alignment = Alignment(horizontal="right")

    last = len(df_sorted) + 1
    tr = last + 1
    ws1.cell(tr, 1, "TOTAL").font = Font(name="Arial", bold=True, color="E6EDF3", size=10)
    ws1.cell(tr, 1).fill = HEADER_FILL
    tc3 = ws1.cell(tr, 3, f"=SUM(C2:C{last})")
    tc3.font = Font(name="Arial", bold=True, color="E6EDF3", size=10)
    tc3.fill = HEADER_FILL; tc3.number_format = '"$"#,##0.0'
    tc3.alignment = Alignment(horizontal="right")
    tc4 = ws1.cell(tr, 4, f"=SUM(D2:D{last})")
    tc4.font = Font(name="Arial", bold=True, color="E6EDF3", size=10)
    tc4.fill = HEADER_FILL; tc4.number_format = '"$"#,##0.00'
    tc4.alignment = Alignment(horizontal="right")

    note_r = tr + 2
    ws1.cell(note_r, 1,
        f"Source: US Treasury MSPD · {agg['summary'].get('extracted_at','')[:10]} "
        f"· {agg.get('raw_count',0)} securities"
    ).font = Font(name="Arial", size=9, italic=True, color="9CA3AF")

    # ── Sheet 2: Monthly Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Monthly Summary")
    monthly = agg.get("monthly", [])
    TYPES = ["bill", "note", "bond", "tips", "frn"]
    types_present = [t for t in TYPES if any(r.get(t, 0) for r in monthly)]

    s2_headers = ["Month"] + [t.upper() + " ($B)" for t in types_present] + ["TOTAL ($B)"]
    s2_widths  = [12] + [14] * len(types_present) + [14]
    for ci, (h, w) in enumerate(zip(s2_headers, s2_widths), 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "B2"

    for ri, rec in enumerate(monthly, 2):
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws2.cell(ri, 1, rec.get("ym", rec.get("maturity_ym", ""))).font = BLUE_FONT
        ws2.cell(ri, 1).fill = fill
        ws2.cell(ri, 1).alignment = Alignment(horizontal="center")

        for ci, t in enumerate(types_present, 2):
            val_m = rec.get(t, 0) or 0
            c = ws2.cell(ri, ci, round(val_m / 1000, 2) if val_m else None)
            c.font = BLUE_FONT; c.fill = fill
            c.number_format = '"$"#,##0.00;"-"'; c.alignment = Alignment(horizontal="right")

        total_col = 2 + len(types_present)
        fl = get_column_letter(2); ll = get_column_letter(1 + len(types_present))
        tc = ws2.cell(ri, total_col, f"=SUM({fl}{ri}:{ll}{ri})")
        tc.font = BLACK_FONT; tc.fill = fill
        tc.number_format = '"$"#,##0.00'; tc.alignment = Alignment(horizontal="right")

    last_data = len(monthly) + 1
    gt = last_data + 1
    ws2.cell(gt, 1, "GRAND TOTAL").font = Font(name="Arial", bold=True, color="E6EDF3", size=10)
    ws2.cell(gt, 1).fill = HEADER_FILL
    ws2.cell(gt, 1).alignment = Alignment(horizontal="center")
    for ci in range(2, 2 + len(types_present) + 1):
        cl = get_column_letter(ci)
        gc = ws2.cell(gt, ci, f"=SUM({cl}2:{cl}{last_data})")
        gc.font = Font(name="Arial", bold=True, color="E6EDF3", size=10)
        gc.fill = HEADER_FILL
        gc.number_format = '"$"#,##0.00'; gc.alignment = Alignment(horizontal="right")

    # Key stats summary box
    s = agg["summary"]
    sc = 2 + len(types_present) + 2
    ws2.cell(2, sc, "Key Stats").font = Font(name="Arial", bold=True, size=10)
    ws2.column_dimensions[get_column_letter(sc)].width = 34
    ws2.column_dimensions[get_column_letter(sc + 1)].width = 14
    stats = [
        ("Next 12-Month Maturities ($B):",   s["total_next_12m_billions"]),
        ("Peak Month:",                       s.get("peak_month", "—")),
        ("Peak Month Amount ($B):",           s.get("peak_month_billions", 0)),
        ("Avg Monthly ($B):",                 s.get("avg_monthly_billions", 0)),
        ("Interest Cost @ 4.5% ($B/yr):",    round(s["total_next_12m_billions"] * 0.045, 1)),
        ("Interest Cost @ 3.5% ($B/yr):",    round(s["total_next_12m_billions"] * 0.035, 1)),
        ("Saving at 3.5% vs 4.5% ($B/yr):",  round(s["total_next_12m_billions"] * 0.01, 1)),
    ]
    for i, (lbl, val) in enumerate(stats, 3):
        ws2.cell(i, sc, lbl).font = Font(name="Arial", size=10, color="6B7280")
        vc = ws2.cell(i, sc + 1, val)
        vc.font = BLUE_FONT; vc.alignment = Alignment(horizontal="right")
        if isinstance(val, float):
            vc.number_format = '"$"#,##0.0'

    wb.save(str(output_path))
    log.info(f"Excel saved → {output_path}")
    return output_path


# ── HTML Dashboard ────────────────────────────────────────────────────────────

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>US Treasury Debt Maturity Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  :root {{
    --bg:#0d1117;--surface:#161b22;--border:#30363d;--text:#e6edf3;
    --muted:#8b949e;--accent:#58a6ff;--red:#f85149;--green:#3fb950;
    --yellow:#d29922;--orange:#f0883e;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;padding:24px}}
  h1{{font-size:1.5rem;font-weight:600;margin-bottom:4px}}
  .sub{{color:var(--muted);font-size:.875rem;margin-bottom:24px}}
  .kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}}
  .kpi{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:16px}}
  .kpi-l{{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px}}
  .kpi-v{{font-size:1.75rem;font-weight:700}}
  .kpi-s{{font-size:.75rem;color:var(--muted);margin-top:4px}}
  .warn{{color:var(--orange)}}.danger{{color:var(--red)}}.ok{{color:var(--green)}}
  .card{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:20px;margin-bottom:20px}}
  .card h2{{font-size:1rem;font-weight:600;margin-bottom:16px}}
  .chart-wrap{{position:relative;height:340px}}
  .filters{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px}}
  .fbtn{{padding:4px 12px;border-radius:20px;border:1px solid var(--border);background:transparent;color:var(--text);cursor:pointer;font-size:.8rem;transition:all .15s}}
  .fbtn.active{{background:var(--accent);border-color:var(--accent);color:#000}}
  table{{width:100%;border-collapse:collapse;font-size:.85rem}}
  th{{text-align:left;padding:8px 12px;color:var(--muted);font-weight:500;font-size:.75rem;text-transform:uppercase;border-bottom:1px solid var(--border)}}
  td{{padding:8px 12px;border-bottom:1px solid var(--border)}}
  tr:hover td{{background:rgba(88,166,255,.05)}}
  .bar{{height:6px;border-radius:3px;background:var(--accent);opacity:.7;display:inline-block;margin-top:4px}}
  footer{{color:var(--muted);font-size:.72rem;margin-top:24px;text-align:center}}
</style>
</head>
<body>
<h1>🏛️ US Treasury — Debt Maturity Schedule</h1>
<p class="sub" id="sub">Loading…</p>
<div class="kpis" id="kpis"></div>
<div class="card">
  <h2>Debt Maturing by Month <span style="color:var(--muted);font-weight:400;font-size:.8rem">($B)</span></h2>
  <div class="filters" id="filters"></div>
  <div class="chart-wrap"><canvas id="mainChart"></canvas></div>
</div>
<div class="card">
  <h2>Refinancing Cost Sensitivity</h2>
  <p style="color:var(--muted);font-size:.8rem;margin-top:-8px;margin-bottom:16px">
    Annual interest if next-12m maturities roll at each yield. Each 1% = <b id="pct1" style="color:var(--orange)">—</b>/yr.
  </p>
  <div class="chart-wrap" style="height:220px"><canvas id="costChart"></canvas></div>
</div>
<div class="card">
  <h2>Monthly Breakdown</h2>
  <table><thead><tr id="thead"></tr></thead><tbody id="tbody"></tbody></table>
</div>
<footer id="foot">treasury_scraper.py</footer>
<script>
const D = {DATA_JSON};
const TYPES = ['bill','note','bond','tips','frn'];
const TC = {{bill:'rgba(240,136,62,.8)',note:'rgba(88,166,255,.8)',bond:'rgba(63,185,80,.8)',tips:'rgba(248,81,73,.7)',frn:'rgba(139,148,158,.6)'}};
const fB = v => v>0 ? '$'+Math.round(v/1000)+'B' : '—';
const s = D.summary;

// KPIs
document.getElementById('sub').textContent =
  'Monthly Statement of Public Debt · '+D.raw_count+' securities · Extracted '+(s.extracted_at||'').slice(0,10);
const total12T = (s.total_next_12m_billions/1000).toFixed(1);
document.getElementById('kpis').innerHTML = [
  ['Maturing Next 12 Months','$'+total12T+'T','Total face value rolling over','danger'],
  ['Peak Month',s.peak_month||'—','$'+s.peak_month_billions+'B maturing','danger'],
  ['Avg Monthly','$'+Math.round(s.avg_monthly_billions)+'B','Over next 12 months','warn'],
  ['Interest Saving (3.5% vs 4.5%)','$'+((s.total_next_12m_billions||0)*0.01).toFixed(0)+'B/yr',
   'At 3.5%: $'+((s.total_next_12m_billions||0)*0.035/1000).toFixed(2)+'T  vs  4.5%: $'+((s.total_next_12m_billions||0)*0.045/1000).toFixed(2)+'T','ok'],
].map(([l,v,sub,cls])=>`<div class="kpi"><div class="kpi-l">${{l}}</div><div class="kpi-v ${{cls}}">${{v}}</div><div class="kpi-s">${{sub}}</div></div>`).join('');

// Main chart
let active = new Set(TYPES), chart;
function buildMain(){{
  const months = D.monthly.map(r=>r.ym||r.maturity_ym);
  const sets = TYPES.filter(t=>D.monthly.some(r=>(r[t]||0)>0)).map(t=>({{
    label:t.toUpperCase(), stack:'s',
    data:D.monthly.map(r=>+((r[t]||0)/1000).toFixed(1)),
    backgroundColor:TC[t]
  }}));
  chart = new Chart(document.getElementById('mainChart').getContext('2d'),{{
    type:'bar', data:{{labels:months,datasets:sets}},
    options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{
        label:c=>` ${{c.dataset.label}}: $${{c.parsed.y.toFixed(0)}}B`,
        footer:ii=>` Total: $${{ii.reduce((s,i)=>s+i.parsed.y,0).toFixed(0)}}B`
      }}}}}},
      scales:{{
        x:{{stacked:true,grid:{{color:'#30363d'}},ticks:{{color:'#8b949e',maxRotation:45}}}},
        y:{{stacked:true,grid:{{color:'#30363d'}},ticks:{{color:'#8b949e',callback:v=>'$'+v+'B'}}}}
      }}
    }}
  }});
  document.getElementById('filters').innerHTML = TYPES.map(t=>
    `<button class="fbtn active" data-t="${{t}}" onclick="tog('${{t}}')">${{t.toUpperCase()}}</button>`
  ).join('');
}}
function tog(t){{
  active.has(t)?active.delete(t):active.add(t);
  document.querySelectorAll('.fbtn').forEach(b=>b.classList.toggle('active',active.has(b.dataset.t)));
  chart.data.datasets.forEach(d=>d.hidden=!active.has(d.label.toLowerCase()));
  chart.update();
}}

// Cost chart
const t12 = (s.total_next_12m_billions||0)*1e9;
document.getElementById('pct1').textContent='$'+((t12*0.01)/1e9).toFixed(0)+'B';
const ys=[2,2.5,3,3.5,4,4.5,5,5.5,6];
new Chart(document.getElementById('costChart').getContext('2d'),{{
  type:'line',
  data:{{labels:ys.map(y=>y.toFixed(1)+'%'),datasets:[{{
    data:ys.map(y=>+(t12*y/100/1e12).toFixed(2)),
    borderColor:'#f85149',backgroundColor:'rgba(248,81,73,.08)',fill:true,tension:.3,
    pointBackgroundColor:ys.map(y=>y<=3.5?'#3fb950':y<=4.5?'#d29922':'#f85149'),pointRadius:5
  }}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},
    tooltip:{{callbacks:{{label:c=>` Cost: $${{c.parsed.y.toFixed(2)}}T/yr`}}}}}},
    scales:{{
      x:{{grid:{{color:'#30363d'}},ticks:{{color:'#8b949e'}}}},
      y:{{grid:{{color:'#30363d'}},ticks:{{color:'#8b949e',callback:v=>'$'+v+'T'}}}}
    }}
  }}
}});

// Table
const types_present = TYPES.filter(t=>D.monthly.some(r=>(r[t]||0)>0));
document.getElementById('thead').innerHTML =
  '<th>Month</th><th>Total</th>'+types_present.map(t=>`<th>${{t.toUpperCase()}}</th>`).join('')+'<th></th>';
const mx = Math.max(...D.monthly.map(r=>r.total_millions||0));
document.getElementById('tbody').innerHTML = D.monthly.map(r=>{{
  const pct=Math.round((r.total_millions||0)/mx*160);
  const col = r.total_millions>mx*.8?'var(--red)':r.total_millions>mx*.5?'var(--orange)':'var(--accent)';
  return `<tr><td><b>${{r.ym||r.maturity_ym}}</b></td>
    <td><b>${{fB(r.total_millions)}}</b></td>
    ${{types_present.map(t=>`<td>${{fB(r[t]||0)}}</td>`).join('')}}
    <td><span class="bar" style="width:${{pct}}px;background:${{col}};opacity:.8"></span></td></tr>`;
}}).join('');

document.getElementById('foot').textContent =
  'treasury_scraper.py · fiscaldata.treasury.gov · '+D.raw_count+' securities';
buildMain();
</script>
</body></html>"""


def generate_dashboard(data: dict, output_path: Path) -> Path:
    html = DASHBOARD_HTML.format(DATA_JSON=json.dumps(data, default=str, indent=2))
    output_path.write_text(html, encoding="utf-8")
    log.info(f"Dashboard saved → {output_path}")
    return output_path


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Treasury MSPD Scraper")
    parser.add_argument("--year",  type=int)
    parser.add_argument("--month", type=int)
    parser.add_argument("--xls",   type=str, help="Path to a local MSPD .xls file")
    parser.add_argument("--out",   type=str, default=str(Path(__file__).parent / "data"),
                        help="Output directory (default: ./data)")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1. Get the XLS file
    if args.xls:
        xls_path = Path(args.xls)
        label = xls_path.stem.replace("MonthlyStatementPublicDebt_Entire_", "")
        if not xls_path.exists():
            log.error(f"File not found: {xls_path}"); sys.exit(1)
    else:
        xls_path, label = find_or_download_xls(args.year, args.month)

    # 2. Extract
    raw_df = extract_from_excel(xls_path)

    # 3. Validate against sheet subtotals
    validate(raw_df, xls_path)

    # 4. Excel first — open this to verify before trusting the JSON
    xlsx_path = out_dir / f"maturity_data_{label}.xlsx"
    export_to_excel(raw_df, aggregate_by_month(raw_df), xlsx_path)

    # 5. Aggregate & save JSON
    agg = aggregate_by_month(raw_df)
    agg["raw_count"] = len(raw_df)
    (out_dir / f"maturity_data_{label}.json").write_text(
        json.dumps(agg, default=str, indent=2), encoding="utf-8")
    (out_dir / "maturity_data_latest.json").write_text(
        json.dumps(agg, default=str, indent=2), encoding="utf-8")
    log.info(f"JSON saved → {out_dir / 'maturity_data_latest.json'}")

    # 6. HTML dashboard
    dash_path = out_dir / f"treasury_dashboard_{label}.html"
    generate_dashboard(agg, dash_path)

    print("\n" + "="*60)
    print("✅  TREASURY SCRAPER COMPLETE")
    print("="*60)
    print(f"  Source:     {xls_path.name}")
    print(f"  Securities: {len(raw_df)}")
    print(f"  Next 12m:   ${agg['summary']['total_next_12m_billions']:.0f}B maturing")
    print(f"  Peak month: {agg['summary']['peak_month']}  "
          f"(${agg['summary']['peak_month_billions']:.0f}B)")
    print(f"\n  Excel  →  {xlsx_path}   ← verify this first")
    print(f"  JSON   →  {out_dir / 'maturity_data_latest.json'}")
    print(f"  Dashboard → {dash_path}")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
